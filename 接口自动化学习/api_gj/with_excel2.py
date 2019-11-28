import openpyxl
from  openpyxl import  Workbook


class with_excel2:

    def __init__(self, path):
        if path:
            self.path = path
            self.wb = self.load_excel()          #读取excel
            self.sheets = self.wb.sheetnames    #获取excel中的所有sheets
            self.ws = self.wb.active             #获取当前表单，默认第一个
        else:
            self.path = 'test.xlsx'
            self.wb = self.load_excel()  # 读取excel
            self.sheets = self.wb.sheetnames  # 获取excel中的所有sheets
            self.ws = self.wb.active  # 获取当前表单，默认第一个
    def load_excel(self):
        workbook = openpyxl.load_workbook(self.path)
        return workbook

    # 写入数据
    def write_data(self, row, col, data1):   #行号为第二行开始，col固定为12
        data = self.ws.cell(row, col, data1)
        self.wb.save(self.path)



# if __name__ == '__main__':
#     aa = with_excel2('test.xlsx')
#     aa.write_data(2, 12, 'pass')

